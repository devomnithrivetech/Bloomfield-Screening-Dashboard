"""Demo screening agent — single Claude LLM call, replaces the full 8-agent pipeline.

Flow per email:
  1. Check Supabase for a cached result — return immediately if found.
  2. Fetch the Gmail message and download all attachments.
  3. Upload raw attachment files to AWS S3 under {gmail_message_id}/.
  4. Extract text from PDFs (pdfplumber) and Excel files (openpyxl).
  5. Call Claude with the analyst+screener system prompt and all document text.
  6. Parse the structured tool-use response.
  7. Populate the Excel screener template (openpyxl), writing Deal Summary
     text directly into the template's styled cells (B8, B13).
  8. Upload the finished Excel to S3 under {gmail_message_id}/Screener_*.xlsx.
  9. Persist the deal record and email record to Supabase.
 10. Return the deal_id (UUID string).
"""
from __future__ import annotations

import asyncio
import base64
import io
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import boto3
import openpyxl
import pdfplumber
from app.core.config import get_settings
from app.core.logging import get_logger
from app.core.token_store import load_token
from app.integrations.claude import get_claude_client
from app.integrations.gmail import credentials_from_token_data, download_attachment, get_message_detail
from app.integrations.supabase import get_supabase

log = get_logger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
#  SYSTEM PROMPT
#  Combines the senior-housing-analyst persona (docs/senior_housing_analyst.md)
#  with the screener output spec (docs/senior_housing_screener.md).
# ─────────────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are an expert senior housing real estate investment analyst with 20+ years of \
experience underwriting IL (Independent Living), AL (Assisted Living), MC (Memory Care), and SNF \
(Skilled Nursing Facility) transactions for institutional bridge and permanent lenders. You have \
personally underwritten hundreds of transactions across multiple market cycles.

YOUR ROLE: Protect capital. You are not a broker or marketer. Be cynical, objective, and \
data-driven. Surface risks and red flags before upsides. If a deal looks too good, it probably is.

OPERATING RULES:
1. MISSING DATA IS A RED FLAG — never invent numbers or make assumptions. Flag every data gap \
explicitly as an underwriting risk. Incomplete documents in senior housing due diligence ARE \
themselves a red flag.
2. VERIFICATION — before finalizing, cross-check every figure against the source documents. \
Correct arithmetic errors silently; note irreconcilable conflicts explicitly.
3. PRECISION — when data is available, reference specific line items; no generalities.

═══════════════════════════════════════════════════════════════════════
ANALYTICAL FRAMEWORK
═══════════════════════════════════════════════════════════════════════

1. FINANCIAL PERFORMANCE & UNDERWRITING

T-12 INCOME STATEMENT — line-by-line review:
• LABOR (Salaries & Wages + Taxes & Benefits + Contract Labor):
  - Industry benchmark: 55–65% of revenue for AL/MC. Flag if outside range.
  - Flag spikes >5% vs. prior periods. Isolate agency/contract labor — high usage = staffing instability.
• MANAGEMENT FEES: Confirm arm's-length vs. related-party. Understated management fees artificially \
inflate NOI (the most common seller tactic to boost perceived value).
• UTILITIES: Spikes indicate deferred maintenance or system failures.
• R&M / PLANT OPERATIONS:
  - CRITICAL FLAG if <$500/unit/yr — sellers suppress this to inflate NOI, leaving the next owner \
the backlog.
  - FLAG if <$1,000/unit/yr. Low R&M is never efficiency; it is deferred capex.
• PRO FORMA vs. T-12: Flag any Pro Forma line >5% better than T-12 without a documented, credible \
explanation (signed rate increase letter, executed contract).

REVENUE AUDIT:
• Separate operating revenue (base rent + care fees) from ancillary (therapy, pharmacy, beauty, \
transport). Ancillary is less predictable and often declines post-acquisition.
• Rate increases + declining occupancy = market positioning problem that capital alone cannot solve.
• PAYOR MIX: Medicaid concentration >70% caps upside, compresses margins, and introduces \
reimbursement risk. Flag it.

STAFFING (PPD Analysis):
• Calculate PPD (hours per patient/resident day) by department: nursing, dietary, housekeeping, \
activities.
• Adequate aggregate PPD can mask dangerously thin overnight coverage — check shift-level data if \
provided.

NOI & CAP RATE:
• Use STABILIZED NOI, never raw T-12 as-is.
• If current occupancy <85% OR occupancy has been declining: do NOT use T-12 as the stabilized \
basis. Apply 10% vacancy assumption; stress-test at 85%.
• Normalize T-12 for: one-time items, below-market management fees, deferred capex reserve \
($350–$600/unit/yr for AL/MC is the standard reserve).
• State the derivation explicitly: Stabilized NOI ÷ Purchase Price = Going-in Cap Rate.

2. MARKET & DEMOGRAPHIC ANALYSIS

Geographic rings: 3-mile, 5-mile, 7-mile radii.
• 75+ AGE COHORT: Current count and 5-year growth projection. This is the primary demand driver \
for needs-based care (AL/MC/SNF).
• HOUSEHOLD INCOME: >$50K/yr for entry-level AL; >$75K/yr for premium private-pay AL/IL. Memory \
care typically requires stronger household wealth.
• OWNER-OCCUPIED HOME VALUES: Proxy for seniors' capacity to self-fund care — markets where \
seniors hold significant home equity support private-pay rate sustainability.

Competitor Set (top 3 within 5 miles for AL/MC; 10 miles for IL):
• Unit count and unit mix, monthly base rent and care fee structure, estimated occupancy, physical \
plant age and recent renovations.
• A subject property priced at a premium to the comp set with declining occupancy has a market \
positioning problem that cannot be solved overnight.

3. OPERATIONAL & REGULATORY RISK

CMS / State Survey History:
• Immediate Jeopardy (IJ) citations are near-automatic PASSes. An IJ means a regulator determined \
a resident was at risk of serious harm or death.
• Infection control and medication management deficiencies carry the highest liability.
• Recurring deficiencies at the same operator = systemic culture issue, not isolated incidents.

Staffing Turnover:
• CNA turnover >100%/yr is severe (industry average: 50–100%+/yr).
• The causal chain: high turnover → higher agency costs → more survey deficiencies → lower resident \
satisfaction → occupancy decline. Turnover is the single most reliable leading indicator of \
operational deterioration.
• Administrator / ED revolving door almost always precedes or accompanies occupancy decline.

4. PHYSICAL PLANT & CAPITAL NEEDS

• Low or declining R&M spend is a deferred capex signal, not operational efficiency.
• HVAC replacement: $15,000–$25,000/unit. For 20+ year assets, budget $5,000–$15,000/unit minimum \
capex reserves absent a formal PCR.
• Long institutional corridors in memory care are a functional problem — modern MC design uses \
small household clusters (10–16 residents), natural light, and circular wandering paths.
• Hotel and apartment conversions often have suboptimal room sizes, inadequate common areas, and \
plumbing configurations that increase operational friction.

═══════════════════════════════════════════════════════════════════════
SCREENER COMPUTATION RULES
═══════════════════════════════════════════════════════════════════════

When data is available, compute these before filling the output:
1.  AL Units = AL Studio + AL 1-Bed + AL 2-Bed
2.  MC Units = MC Companion + MC Studio + MC 1-Bed
3.  Total Units = sum of all unit counts
4.  Total Beds = sum of all bed counts
5.  Bed Occupancy % = weighted average occupancy across all unit types
6.  Rent + Care ($/res/mo) = Rent + Care for each unit type
7.  Rent Monthly Total = Units × Occupancy % × Monthly Rent
8.  Rent + Care Monthly Total = Units × Occupancy % × (Rent + Care)
9.  Year N Proforma Value = Year N NOI ÷ Cap Rate
10. Historical NOI = Total Revenue − Total Expenses
11. DSCR = NOI ÷ Annual Debt Service (Loan × Interest Rate)
12. RevPOR = Total Revenue ÷ (Occupied Units × 12)  [monthly, per occupied unit]

═══════════════════════════════════════════════════════════════════════
DEAL SUMMARY NARRATIVE GUIDELINES
═══════════════════════════════════════════════════════════════════════

Write three narrative sections for the Deal Summary section of the Bloomfield Origination Screener.

PROPERTY OVERVIEW (1–2 paragraphs):
- Property name, location (city, state, county, MSA if identifiable), asset type, unit count and mix.
- Licensing framework (AL/MC/IL licensing status; CMS certification if SNF).
- Current occupancy and census vs. licensed capacity.
- T-12 revenue and reported NOI (actual numbers from documents).
- Normalized NOI (after management fee adjustment, deferred capex reserve addition, one-time item \
removal) and implied value range at market cap rates (e.g., "Normalized NOI of $X implies a value \
of $Y–$Z at 8–9% cap rates").
- Deal type context (acquisition, refi, bridge, construction) and loan request amount.

INVESTMENT HIGHLIGHTS (1 paragraph — lead with what is genuinely strong):
- Census stability and ADC trend: monthly range and full-year average (must reference actual \
numbers from the documents).
- Revenue trajectory YoY (state actual % change from the financials).
- Rent roll annualized vs. T-12 run-rate if calculable.
- Payor diversification or private-pay concentration (whichever is favorable — state actual %).
- Market / demographic tailwinds: 75+ population growth, income profile, home values (reference \
actual data if provided).
- Operator track record: years in business, portfolio size, comparable assets (if available).
- Upside thesis: rate growth headroom vs. comp set, occupancy improvement path, specific expense \
normalization opportunities identified in the financials.

INVESTMENT RISKS & UNDERWRITING FLAGS (1–2 paragraphs — be blunt; do not bury risks):
- Medicaid / MCO concentration and rate ceiling risk (state actual %; flag if >70%).
- Payor mix trend YoY: private-pay conversion or deterioration (quantify if possible).
- R&M / deferred capex exposure: state actual $/unit/yr from T-12; CRITICAL if <$500, FLAG if \
<$1,000.
- Payroll ratio vs. 55–65% ALF benchmark: state actual ratio from T-12.
- Insurance anomalies, real estate tax normalization issues (large credits, year-end true-ups).
- Missing due diligence: surveys, PCR/PCA, staffing/PPD data, licensed capacity documentation, \
purchase price — be specific about what is absent.
- Operator track record risks: new operator (<2 years), single reference asset, thin management depth.
- Overall risk rating: Low / Moderate / Moderate-High / High, with one sentence explaining the \
primary driver.
- Bloomfield recommendation: Proceed / Negotiate / Pass + one sentence rationale that references \
the single most critical risk or opportunity.

═══════════════════════════════════════════════════════════════════════
KEY METRICS FOR THE DEAL DISPLAY GRID
═══════════════════════════════════════════════════════════════════════

Include 8–12 key metrics in key_metrics. Required metrics (include all that have data):
  Purchase Price | Requested Loan (LTV %) | T-12 Revenue | T-12 NOI (EBITDA margin %)
  Year 1 Proforma NOI (% change vs T-12) | Current Occupancy % | Going-in Cap Rate
  DSCR (at specified interest rate) | RevPOR — Revenue per Occupied Resident/month (T-12)
  Labor/Payroll Ratio (% of revenue) | CapEx Reserve ($/unit/yr from T-12 R&M + Plant Ops)
  In-place DY (= T-12 NOI ÷ Requested Loan Amount; express as a percentage)

For each metric, also supply per_unit where the value can be meaningfully expressed on a \
per-unit basis (e.g. Requested Loan → loan per unit, T-12 Revenue → revenue per unit, \
T-12 NOI → NOI per unit, CapEx Reserve → already $/unit). Set per_unit to null for metrics \
where a per-unit figure is not meaningful (e.g. Occupancy %, DSCR, cap rates, ratios).

Flag as "warn" if:
  LTV >75% | DSCR <1.20x | Occupancy <85% | T-12 NOI margin <15% | Pro Forma NOI >T-12 by >20%
  Cap rate <7% | Payroll ratio outside 55–65% | R&M <$1,000/unit/yr | In-place DY <8%

Flag as "ok" if within acceptable range.

═══════════════════════════════════════════════════════════════════════
ADDITIONAL EXTRACTION REQUIREMENTS
═══════════════════════════════════════════════════════════════════════

FINANCIAL SUMMARY (financial_summary array in screening_result):
Produce exactly these four rows in order, formatting dollar values with $ and commas:
  label: "T-12 NOI"            | value: T-12 NOI dollar amount
  label: "T-3M Annualized NOI" | value: T-3M annualized NOI dollar amount
  label: "Pro Forma NOI"       | value: Year 1 proforma NOI dollar amount
  label: "Current Occupancy"   | value: occupancy percentage string (e.g. "87.3%")
For the first three rows, compute dy = (NOI ÷ Requested Loan) expressed as a percentage \
string (e.g. "9.2%"). Set dy to null for Current Occupancy. If a value is unavailable, \
use "N/A" for value and null for dy.

SOURCES & USES (sources_and_uses object in screening_result):
Extract the Sources & Uses table from the financing memo if present. If not explicitly \
provided, reconstruct it from the loan terms data. Typical structure:
  Sources: Loan Proceeds (net of reserves), Interest Reserve, Equity / Sponsor Contribution
  Uses: Purchase Price (or payoff amount), CapEx Reserve, Origination Fee, \
        Interest Reserve, Closing Costs, Other
Format every row with: item (string), total (dollar string with $ and commas), \
per_unit (dollar/unit string), pct (percentage of total, e.g. "72.4%"). \
If no data is available for this section, omit the field entirely.

SPONSOR OVERVIEW (sponsor_overview in deal_summary):
Write 1–2 sentences on the proposed borrower/sponsor entity. Include entity name, \
operator experience, years in business, portfolio size, and geographic focus if available. \
If sponsor information is absent from the documents, write a single sentence noting \
what information was not provided.

LOCATION SUMMARY (location_summary in deal_summary):
Write 1–2 sentences on the property's market and location context. Reference the MSA, \
75+ population size or growth trajectory, household income levels, home values, and \
proximity to healthcare infrastructure where data is available. Note any data gaps.

CALL submit_screening_result once with all extracted and computed data. Output nothing outside the \
tool call."""


# ─────────────────────────────────────────────────────────────────────────────
#  TOOL SCHEMA  (forces structured JSON output from Claude)
# ─────────────────────────────────────────────────────────────────────────────
_FIN_PERIOD: dict[str, Any] = {
    "type": "object",
    "properties": {
        "il_revenue":                 {"type": ["number", "null"]},
        "al_revenue":                 {"type": ["number", "null"]},
        "mc_revenue":                 {"type": ["number", "null"]},
        "community_fee":              {"type": ["number", "null"]},
        "additional_occupant_income": {"type": ["number", "null"]},
        "other_income":               {"type": ["number", "null"]},
        "salaries_wages":             {"type": ["number", "null"]},
        "taxes_benefits":             {"type": ["number", "null"]},
        "contract_labor":             {"type": ["number", "null"]},
        "general_administrative":     {"type": ["number", "null"]},
        "plant_operations":           {"type": ["number", "null"]},
        "insurance":                  {"type": ["number", "null"]},
        "real_estate_taxes":          {"type": ["number", "null"]},
        "housekeeping_laundry":       {"type": ["number", "null"]},
        "activities_social_services": {"type": ["number", "null"]},
        "resident_care":              {"type": ["number", "null"]},
        "utilities_fixed":            {"type": ["number", "null"]},
        "utilities_variable":         {"type": ["number", "null"]},
        "dietary":                    {"type": ["number", "null"]},
        "food_costs":                 {"type": ["number", "null"]},
        "marketing":                  {"type": ["number", "null"]},
        "bad_debt":                   {"type": ["number", "null"]},
        "commissions":                {"type": ["number", "null"]},
        "management_fee":             {"type": ["number", "null"]},
    },
}

_PROFORMA_YEAR: dict[str, Any] = {
    "type": "object",
    "properties": {
        "noi":      {"type": ["number", "null"]},
        "cap_rate": {"type": "number"},
    },
}

SCREENER_TOOL: dict[str, Any] = {
    "name": "submit_screening_result",
    "description": (
        "Submit the complete Bloomfield Capital Origination Screener result for a senior housing deal. "
        "Call this once with all extracted, computed, and narrative data after completing your analysis."
    ),
    "input_schema": {
        "type": "object",
        "required": ["property_info", "deal_summary", "screening_result"],
        "properties": {
            "property_info": {
                "type": "object",
                "required": ["property_name", "deal_type", "property_type"],
                "properties": {
                    "property_name":    {"type": "string"},
                    "address":          {"type": ["string", "null"]},
                    "city_state":       {"type": ["string", "null"]},
                    "county":           {"type": ["string", "null"]},
                    "deal_type":        {"type": "string", "description": "Bridge, Construction, Permanent, etc."},
                    "property_type":    {"type": "string", "description": "AL/MC, IL/AL/MC, AL Only, etc."},
                    "year_built":       {"type": ["integer", "null"]},
                    "building_sf":      {"type": ["number", "null"]},
                    "al_units":         {"type": ["integer", "null"]},
                    "mc_units":         {"type": ["integer", "null"]},
                    "il_units":         {"type": ["integer", "null"]},
                    "total_units":      {"type": ["integer", "null"]},
                    "total_beds":       {"type": ["integer", "null"]},
                    "bed_occupancy_pct":{"type": ["number", "null"]},
                },
            },
            "unit_mix": {
                "type": "array",
                "items": {
                    "type": "object",
                    "required": ["unit_type"],
                    "properties": {
                        "unit_type": {
                            "type": "string",
                            "enum": [
                                "AL Studio", "AL 1-Bed", "AL 2-Bed",
                                "MC Companion", "MC Studio", "MC 1-Bed",
                                "IL Studio", "IL 1-Bed", "IL 2-Bed",
                            ],
                        },
                        "total_units":          {"type": ["integer", "null"]},
                        "total_beds":           {"type": ["integer", "null"]},
                        "occupancy_pct":        {"type": ["number", "null"]},
                        "rent_per_resident_mo": {"type": ["number", "null"]},
                        "care_per_resident_mo": {"type": ["number", "null"]},
                    },
                },
            },
            "loan_terms": {
                "type": "object",
                "properties": {
                    "requested_loan_amount":  {"type": ["number", "null"]},
                    "purchase_price":         {"type": ["number", "null"]},
                    "interest_rate":          {"type": "number", "description": "Decimal, e.g. 0.11 for 11%"},
                    "interest_reserve_months":{"type": "integer"},
                    "capex_reserve":          {"type": "number"},
                    "bloomfield_points_pct":  {"type": "number"},
                    "broker_points_pct":      {"type": "number"},
                },
            },
            "financials": {
                "type": "object",
                "description": "Annual financial data keyed by period. Omit periods for which no data exists.",
                "properties": {
                    "ye_2022":        _FIN_PERIOD,
                    "ye_2023":        _FIN_PERIOD,
                    "ye_2024":        _FIN_PERIOD,
                    "t11m_annualized":_FIN_PERIOD,
                    "t3m_annualized": _FIN_PERIOD,
                },
            },
            "proforma": {
                "type": "object",
                "properties": {
                    "year_1": _PROFORMA_YEAR,
                    "year_2": _PROFORMA_YEAR,
                    "year_3": _PROFORMA_YEAR,
                },
            },
            "deal_summary": {
                "type": "object",
                "required": ["property_overview", "investment_highlights", "investment_risks"],
                "properties": {
                    "property_overview":     {"type": "string"},
                    "investment_highlights": {"type": "string"},
                    "investment_risks":      {"type": "string"},
                    "sponsor_overview":      {"type": ["string", "null"]},
                    "location_summary":      {"type": ["string", "null"]},
                },
            },
            "screening_result": {
                "type": "object",
                "required": ["recommendation", "risk_rating", "key_metrics"],
                "properties": {
                    "recommendation": {
                        "type": "string",
                        "enum": ["proceed", "negotiate", "pass"],
                    },
                    "risk_rating": {
                        "type": "string",
                        "enum": ["low", "moderate", "moderate_high", "high"],
                    },
                    "confidence": {
                        "type": "number",
                        "description": "0.0–1.0 based on data completeness and analytical certainty",
                    },
                    "key_metrics": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "required": ["label", "value", "flag"],
                            "properties": {
                                "label":    {"type": "string"},
                                "value":    {"type": "string"},
                                "flag":     {"type": "string", "enum": ["ok", "warn"]},
                                "per_unit": {"type": ["string", "null"]},
                            },
                        },
                    },
                    "financial_summary": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "required": ["label", "value"],
                            "properties": {
                                "label": {"type": "string"},
                                "value": {"type": "string"},
                                "dy":    {"type": ["string", "null"]},
                            },
                        },
                    },
                    "sources_and_uses": {
                        "type": "object",
                        "properties": {
                            "sources": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "required": ["item", "total", "per_unit", "pct"],
                                    "properties": {
                                        "item":     {"type": "string"},
                                        "total":    {"type": "string"},
                                        "per_unit": {"type": "string"},
                                        "pct":      {"type": "string"},
                                    },
                                },
                            },
                            "uses": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "required": ["item", "total", "per_unit", "pct"],
                                    "properties": {
                                        "item":     {"type": "string"},
                                        "total":    {"type": "string"},
                                        "per_unit": {"type": "string"},
                                        "pct":      {"type": "string"},
                                    },
                                },
                            },
                        },
                    },
                },
            },
        },
    },
}


# ─────────────────────────────────────────────────────────────────────────────
#  PUBLIC ENTRY POINTS
# ─────────────────────────────────────────────────────────────────────────────

async def run_manual_upload_screening(
    user_id: str,
    email_id: str,
    screened_email_id: str,
    raw_attachments: list[dict[str, Any]],
    email_meta: dict[str, Any],
) -> str:
    """Process a manually-uploaded deal and return the deal_id (UUID string).

    Skips all Gmail I/O — attachments are already in memory from the upload
    request.  Uses screened_email_id (the PK of the screened_emails row) as the
    pipeline correlation key instead of gmail_message_id.
    """
    settings = get_settings()
    supabase = get_supabase()
    loop = asyncio.get_event_loop()

    # email_id (internal UUID) doubles as the S3 folder prefix for uploads
    s3_folder = email_id

    # Initialise pipeline stages on the pre-created screened_emails stub
    _init_manual_screened_email(supabase, screened_email_id)

    try:
        # Advance: email_received → complete, parsing_attachments → in_progress
        _advance_stage_by_id(
            supabase, screened_email_id,
            new_status="parsing_attachments",
            completing="email_received",
            next_stage="parsing_attachments",
        )

        # Upload raw attachments to S3 (best-effort — pipeline continues on failure)
        try:
            s3_attachment_keys = await loop.run_in_executor(
                None, _upload_attachments_sync, s3_folder, raw_attachments, settings
            )
        except Exception as exc:
            log.warning("manual_upload_s3_skipped", error=str(exc))
            s3_attachment_keys = []

        # Record S3 keys on the emails row so the row is queryable later
        try:
            supabase.table("emails").update({
                "attachments": [
                    {"filename": k["filename"], "s3_key": k["s3_key"], "type": k["type"]}
                    for k in s3_attachment_keys
                ]
            }).eq("id", email_id).execute()
        except Exception as exc:
            log.warning("manual_upload_attachments_update_failed", error=str(exc))

        # Extract text from PDFs / Excel files
        doc_text_blocks = _build_document_text_blocks(raw_attachments)

        _advance_stage_by_id(
            supabase, screened_email_id,
            new_status="extracting_financials",
            completing="parsing_attachments",
            next_stage="extracting_financials",
        )

        # Build a lightweight email_detail object for _call_claude
        from types import SimpleNamespace
        email_detail = SimpleNamespace(
            sender=email_meta.get("sender") or "Manual Upload",
            sender_email=email_meta.get("sender_email"),
            subject=email_meta.get("subject") or "Manual Upload",
            received_at=datetime.now(timezone.utc),
            body_text=email_meta.get("body_text"),
            attachments=[],
        )

        log.info("manual_upload_calling_claude", email_id=email_id)
        result = await _call_claude(email_detail, doc_text_blocks, settings, additional_instructions=None)

        _advance_stage_by_id(
            supabase, screened_email_id,
            new_status="running_screener",
            completing="extracting_financials",
            next_stage="running_screener",
        )

        # Generate Excel screener
        deal_id = str(uuid.uuid4())
        screener_s3_key: str | None = None
        try:
            excel_bytes = _generate_excel_screener(result, settings.screener_template_path)
            prop_name = (result.get("property_info") or {}).get("property_name", "deal")
            safe_name = re.sub(r"[^\w\-]", "_", prop_name).strip("_")[:50]
            screener_s3_key = f"{s3_folder}/Screener_{safe_name}.xlsx"
            await loop.run_in_executor(
                None, _s3_put_bytes,
                screener_s3_key, excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                settings,
            )
            log.info("manual_upload_screener_uploaded", s3_key=screener_s3_key)
        except Exception as exc:
            log.warning("manual_upload_screener_failed", error=str(exc))

        # Persist deal + mark screened_emails complete
        _persist_manual_upload_to_supabase(
            supabase, user_id, deal_id, email_id, screened_email_id,
            result, screener_s3_key, s3_attachment_keys, email_detail, s3_folder,
        )
        log.info("manual_upload_screening_complete", deal_id=deal_id)
        return deal_id

    except Exception as exc:
        log.error("manual_upload_screening_failed", email_id=email_id, error=str(exc))
        _fail_screened_email_by_id(supabase, screened_email_id, str(exc))
        try:
            supabase.table("emails").update({"status": "failed"}).eq("id", email_id).execute()
        except Exception:
            pass
        raise


async def run_demo_screening(
    user_id: str,
    gmail_message_id: str,
    extra_attachments: list[dict[str, Any]] | None = None,
    additional_instructions: str | None = None,
) -> str:
    """Process a deal email end-to-end and return the deal_id (UUID string).

    On a second call for the same (user_id, gmail_message_id) pair the cached
    deal_id is returned immediately without re-processing.

    `extra_attachments` — optional list of {filename, type, data} dicts uploaded
    by the user from their device alongside the email.  Stored in S3 under the
    same gmail_message_id folder and fed to the LLM as additional documents.

    `additional_instructions` — optional free-text analyst instructions injected
    verbatim into the screener prompt before the final analysis task.
    """
    settings = get_settings()
    supabase = get_supabase()

    # 1. Return cached result if already processed
    cached = _check_cached_result(supabase, user_id, gmail_message_id)
    if cached:
        log.info("demo_cache_hit", gmail_message_id=gmail_message_id, deal_id=cached)
        _sync_cached_screened_email(supabase, user_id, gmail_message_id, cached)
        return cached

    # 2. Load Gmail credentials
    token_data = await load_token(user_id)
    if not token_data:
        raise RuntimeError("Gmail not connected — cannot fetch email for processing")
    creds = credentials_from_token_data(token_data)

    try:
        # 3. Fetch full email from Gmail
        email_detail = await get_message_detail(creds, gmail_message_id)
        if email_detail is None:
            raise RuntimeError(f"Gmail message {gmail_message_id} not found")

        # Writes status=processing to emails table and inits pipeline stages on
        # the screened_emails row (email_received: in_progress, rest: pending).
        _mark_email_processing(supabase, user_id, gmail_message_id, email_detail)

        # 4. Download attachments  (email_received → complete, parsing_attachments → in_progress)
        _advance_screened_email_stage(
            supabase, user_id, gmail_message_id,
            new_status="parsing_attachments",
            completing="email_received",
            next_stage="parsing_attachments",
        )
        loop = asyncio.get_event_loop()
        raw_attachments: list[dict[str, Any]] = []
        for att in email_detail.attachments:
            if att.id:
                try:
                    data = await download_attachment(creds, gmail_message_id, att.id)
                    raw_attachments.append({
                        "id": att.id,
                        "filename": att.filename,
                        "type": att.type.value,
                        "data": data,
                    })
                    log.info("attachment_downloaded", filename=att.filename, bytes=len(data))
                except Exception as exc:
                    log.warning("attachment_download_failed", filename=att.filename, error=str(exc))

        # 5. Upload Gmail attachments to S3 (best-effort — pipeline continues on failure)
        try:
            s3_attachment_keys = await loop.run_in_executor(
                None, _upload_attachments_sync, gmail_message_id, raw_attachments, settings
            )
        except Exception as exc:
            log.warning("s3_attachments_skipped", error=str(exc))
            s3_attachment_keys = []

        # 5b. Upload user-supplied extra attachments to the same S3 folder
        if extra_attachments:
            try:
                extra_s3_keys = await loop.run_in_executor(
                    None, _upload_attachments_sync, gmail_message_id, extra_attachments, settings
                )
                s3_attachment_keys.extend(extra_s3_keys)
                log.info(
                    "extra_attachments_uploaded",
                    count=len(extra_s3_keys),
                    gmail_message_id=gmail_message_id,
                )
            except Exception as exc:
                log.warning("s3_extra_attachments_skipped", error=str(exc))

        # 6. Extract document text from all attachments
        #    (parsing_attachments → complete, extracting_financials → in_progress)
        doc_text_blocks = _build_document_text_blocks(raw_attachments)
        if extra_attachments:
            doc_text_blocks.extend(_build_document_text_blocks(extra_attachments))

        _advance_screened_email_stage(
            supabase, user_id, gmail_message_id,
            new_status="extracting_financials",
            completing="parsing_attachments",
            next_stage="extracting_financials",
        )

        # 7. Call Claude with the full analyst prompt
        log.info("demo_calling_claude", gmail_message_id=gmail_message_id)
        result = await _call_claude(
            email_detail, doc_text_blocks, settings,
            additional_instructions=additional_instructions,
        )

        # (extracting_financials → complete, running_screener → in_progress)
        _advance_screened_email_stage(
            supabase, user_id, gmail_message_id,
            new_status="running_screener",
            completing="extracting_financials",
            next_stage="running_screener",
        )

        # 8. Generate Excel screener
        deal_id = str(uuid.uuid4())
        screener_s3_key: str | None = None
        try:
            excel_bytes = _generate_excel_screener(result, settings.screener_template_path)
            prop_name = (result.get("property_info") or {}).get("property_name", "deal")
            safe_name = re.sub(r"[^\w\-]", "_", prop_name).strip("_")[:50]
            screener_s3_key = f"{gmail_message_id}/Screener_{safe_name}.xlsx"
            await loop.run_in_executor(
                None, _s3_put_bytes,
                screener_s3_key, excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                settings,
            )
            log.info("screener_uploaded", s3_key=screener_s3_key)
        except Exception as exc:
            log.warning("screener_generation_failed", error=str(exc))

        # 9. Persist to Supabase (calls _complete_screened_email internally)
        _persist_to_supabase(
            supabase, user_id, deal_id, gmail_message_id,
            result, screener_s3_key, s3_attachment_keys, email_detail,
        )
        log.info("demo_screening_complete", deal_id=deal_id)
        return deal_id

    except Exception as exc:
        log.error("demo_screening_failed", gmail_message_id=gmail_message_id, error=str(exc))
        _fail_screened_email(supabase, user_id, gmail_message_id, str(exc))
        raise


# ─────────────────────────────────────────────────────────────────────────────
#  SUPABASE HELPERS — manual-upload variants (lookup by screened_email_id PK)
# ─────────────────────────────────────────────────────────────────────────────

def _init_manual_screened_email(supabase: Any, screened_email_id: str) -> None:
    """Set initial pipeline stages on a pre-created screened_emails row."""
    now = datetime.now(timezone.utc).isoformat()
    initial_pipeline = [
        {"stage": "email_received",        "status": "in_progress", "started_at": now,  "finished_at": None},
        {"stage": "parsing_attachments",   "status": "pending",     "started_at": None, "finished_at": None},
        {"stage": "extracting_financials", "status": "pending",     "started_at": None, "finished_at": None},
        {"stage": "running_screener",      "status": "pending",     "started_at": None, "finished_at": None},
        {"stage": "complete",              "status": "pending",     "started_at": None, "finished_at": None},
    ]
    try:
        supabase.table("screened_emails").update({
            "processing_status": "email_received",
            "pipeline":          initial_pipeline,
        }).eq("id", screened_email_id).execute()
    except Exception as exc:
        log.warning("init_manual_screened_email_failed", screened_email_id=screened_email_id, error=str(exc))


def _advance_stage_by_id(
    supabase: Any,
    screened_email_id: str,
    new_status: str,
    completing: str,
    next_stage: str | None = None,
) -> None:
    """Like _advance_screened_email_stage but keyed on the row's UUID PK."""
    now = datetime.now(timezone.utc).isoformat()
    try:
        resp = (
            supabase.table("screened_emails")
            .select("pipeline")
            .eq("id", screened_email_id)
            .limit(1)
            .execute()
        )
        if not resp.data:
            return
        pipeline: list[dict] = list(resp.data[0].get("pipeline") or [])
        updated = []
        for s in pipeline:
            s = dict(s)
            if s["stage"] == completing:
                s["status"] = "completed"
                s["finished_at"] = now
            elif next_stage and s["stage"] == next_stage:
                s["status"] = "in_progress"
                s["started_at"] = now
            updated.append(s)
        supabase.table("screened_emails").update(
            {"pipeline": updated, "processing_status": new_status}
        ).eq("id", screened_email_id).execute()
    except Exception as exc:
        log.warning("advance_stage_by_id_failed", completing=completing, error=str(exc))


def _fail_screened_email_by_id(
    supabase: Any,
    screened_email_id: str,
    error_msg: str,
) -> None:
    """Mark the screened_emails row as failed, keyed on its UUID PK."""
    now = datetime.now(timezone.utc).isoformat()
    try:
        resp = (
            supabase.table("screened_emails")
            .select("pipeline")
            .eq("id", screened_email_id)
            .limit(1)
            .execute()
        )
        if not resp.data:
            return
        pipeline: list[dict] = list(resp.data[0].get("pipeline") or [])
        updated = []
        for s in pipeline:
            s = dict(s)
            if s.get("status") == "in_progress":
                s["status"] = "failed"
                s["finished_at"] = now
                s["detail"] = error_msg[:200]
            updated.append(s)
        supabase.table("screened_emails").update(
            {"pipeline": updated, "processing_status": "failed"}
        ).eq("id", screened_email_id).execute()
    except Exception as exc:
        log.warning("fail_screened_email_by_id_failed", error=str(exc))


def _complete_screened_email_by_id(
    supabase: Any,
    screened_email_id: str,
    deal_id: str,
    screened_title: str | None,
    screener_s3_key: str | None,
    pipeline_stages: list[dict],
) -> None:
    """Mark the screened_emails row as complete, keyed on its UUID PK."""
    update: dict[str, Any] = {
        "processing_status": "complete",
        "deal_id":           deal_id,
        "screened_title":    screened_title,
        "screener_s3_key":   screener_s3_key,
        "pipeline":          pipeline_stages,
    }
    try:
        supabase.table("screened_emails").update(update).eq("id", screened_email_id).execute()
    except Exception as exc:
        log.warning("complete_screened_email_by_id_failed", error=str(exc))


def _persist_manual_upload_to_supabase(
    supabase: Any,
    user_id: str,
    deal_id: str,
    email_id: str,
    screened_email_id: str,
    result: dict[str, Any],
    screener_s3_key: str | None,
    s3_attachment_keys: list[dict],
    email_detail: Any,
    s3_folder: str,
) -> None:
    """Write the deal row and mark both emails + screened_emails as complete."""
    prop     = result.get("property_info") or {}
    summary  = result.get("deal_summary") or {}
    screening = result.get("screening_result") or {}
    now = datetime.now(timezone.utc).isoformat()

    pipeline_stages = [
        {"stage": "email_received",        "status": "completed", "finished_at": now},
        {"stage": "parsing_attachments",   "status": "completed", "finished_at": now},
        {"stage": "extracting_financials", "status": "completed", "finished_at": now},
        {"stage": "running_screener",      "status": "completed", "finished_at": now},
        {"stage": "complete",              "status": "completed", "finished_at": now},
    ]

    # Finalise the emails row
    try:
        supabase.table("emails").update({
            "status":      "processed",
            "deal_id":     deal_id,
            "attachments": [
                {"filename": k["filename"], "s3_key": k["s3_key"], "type": k["type"]}
                for k in s3_attachment_keys
            ],
        }).eq("id", email_id).execute()
    except Exception as exc:
        log.warning("manual_email_finalise_failed", email_id=email_id, error=str(exc))

    # Insert deal row (mirrors _persist_to_supabase structure exactly)
    deal_row: dict[str, Any] = {
        "id":                    deal_id,
        "user_id":               user_id,
        "email_id":              email_id,
        "property_name":         prop.get("property_name", "Unknown Property"),
        "asset_class":           prop.get("property_type"),
        "recommendation":        screening.get("recommendation"),
        "confidence":            screening.get("confidence"),
        "risk_rating":           screening.get("risk_rating"),
        "screener_storage_path": screener_s3_key,
        "email_draft":           _build_email_draft(result),
        "metrics":               screening.get("key_metrics", []),
        "key_metrics":           screening.get("key_metrics", []),
        "highlights": [
            {
                "title":  "Investment Highlights",
                "detail": summary.get("investment_highlights", ""),
            }
        ],
        "risks": [
            {
                "title":    "Investment Risks & Underwriting Flags",
                "detail":   summary.get("investment_risks", ""),
                "severity": screening.get("risk_rating", "moderate"),
            }
        ],
        "pipeline":          pipeline_stages,
        "financial_summary": screening.get("financial_summary"),
        "sources_and_uses":  screening.get("sources_and_uses"),
        "sponsor_overview":  summary.get("sponsor_overview"),
        "location_summary":  summary.get("location_summary"),
        "property_info":     result.get("property_info") or {},
        "financials": {
            "historical": result.get("financials") or {},
            "proforma":   result.get("proforma") or {},
        },
        "s3_folder_key": s3_folder,
        "created_at":    now,
        "updated_at":    now,
    }
    try:
        supabase.table("deals").insert(deal_row).execute()
    except Exception as exc:
        log.error("manual_deal_insert_failed", deal_id=deal_id, error=str(exc))
        raise

    _complete_screened_email_by_id(
        supabase, screened_email_id, deal_id,
        screened_title=prop.get("property_name"),
        screener_s3_key=screener_s3_key,
        pipeline_stages=pipeline_stages,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  SUPABASE HELPERS — Gmail flow (lookup by gmail_message_id)
# ─────────────────────────────────────────────────────────────────────────────

def _find_email_row_id(supabase: Any, user_id: str, gmail_message_id: str) -> str | None:
    """Return the internal uuid of an existing email row, or None."""
    resp = (
        supabase.table("emails")
        .select("id")
        .eq("user_id", user_id)
        .eq("gmail_message_id", gmail_message_id)
        .limit(1)
        .execute()
    )
    return str(resp.data[0]["id"]) if resp.data else None


def _safe_email_write(
    supabase: Any,
    user_id: str,
    gmail_message_id: str,
    row: dict[str, Any],
) -> str | None:
    """Insert or update an email row without relying on ON CONFLICT constraints.

    PostgREST's upsert(on_conflict=...) requires a matching full UNIQUE
    constraint.  Migration 0004 only created a PARTIAL unique index, which
    PostgreSQL rejects with error 42P10.  This helper avoids that dependency:
    it first checks whether the row exists, then inserts or updates accordingly.

    Returns the internal row id on success, or None on failure.
    """
    try:
        existing_id = _find_email_row_id(supabase, user_id, gmail_message_id)
        if existing_id:
            supabase.table("emails").update(row).eq("id", existing_id).execute()
            return existing_id
        else:
            resp = supabase.table("emails").insert(row).execute()
            return str(resp.data[0]["id"]) if resp.data else None
    except Exception as exc:
        log.warning("email_write_failed", gmail_message_id=gmail_message_id, error=str(exc))
        return None


def _mark_email_processing(
    supabase: Any,
    user_id: str,
    gmail_message_id: str,
    email_detail: Any,
) -> None:
    """Write status='processing' before the LLM call so that a page refresh
    during the pipeline still shows the in-flight state on the dashboard."""
    row: dict[str, Any] = {
        "user_id":          user_id,
        "gmail_message_id": gmail_message_id,
        "sender":           email_detail.sender or "Unknown",
        "subject":          email_detail.subject or "(no subject)",
        "body_text":        email_detail.body_text,
        "received_at":      (
            email_detail.received_at.isoformat()
            if email_detail.received_at
            else datetime.now(timezone.utc).isoformat()
        ),
        "status":           "processing",
        "attachments":      [],
    }
    _safe_email_write(supabase, user_id, gmail_message_id, row)
    _upsert_screened_email_queued(supabase, user_id, gmail_message_id, email_detail)


def _upsert_screened_email_queued(
    supabase: Any,
    user_id: str,
    gmail_message_id: str,
    email_detail: Any,
) -> None:
    """Insert a screened_emails row when processing starts (idempotent)."""
    now = datetime.now(timezone.utc).isoformat()
    initial_pipeline = [
        {"stage": "email_received",        "status": "in_progress", "started_at": now,  "finished_at": None},
        {"stage": "parsing_attachments",   "status": "pending",     "started_at": None, "finished_at": None},
        {"stage": "extracting_financials", "status": "pending",     "started_at": None, "finished_at": None},
        {"stage": "running_screener",      "status": "pending",     "started_at": None, "finished_at": None},
        {"stage": "complete",              "status": "pending",     "started_at": None, "finished_at": None},
    ]
    row: dict[str, Any] = {
        "user_id":               user_id,
        "gmail_message_id":      gmail_message_id,
        "subject":               email_detail.subject or "(no subject)",
        "sender":                email_detail.sender or "Unknown",
        "sender_email":          getattr(email_detail, "sender_email", None),
        "received_at":           (
            email_detail.received_at.isoformat()
            if email_detail.received_at
            else now
        ),
        "sent_for_screening_at": now,
        "processing_status":     "email_received",
        "pipeline":              initial_pipeline,
    }
    try:
        existing = (
            supabase.table("screened_emails")
            .select("id")
            .eq("user_id", user_id)
            .eq("gmail_message_id", gmail_message_id)
            .limit(1)
            .execute()
        )
        if existing.data:
            supabase.table("screened_emails").update(row).eq("id", existing.data[0]["id"]).execute()
        else:
            supabase.table("screened_emails").insert(row).execute()
    except Exception as exc:
        log.warning("screened_email_upsert_failed", gmail_message_id=gmail_message_id, error=str(exc))


def _complete_screened_email(
    supabase: Any,
    user_id: str,
    gmail_message_id: str,
    deal_id: str,
    screened_title: str | None,
    screener_s3_key: str | None,
    pipeline_stages: list[dict],
) -> None:
    """Mark the screened_emails row as complete once the deal is persisted."""
    update: dict[str, Any] = {
        "processing_status": "complete",
        "deal_id":           deal_id,
        "screened_title":    screened_title,
        "screener_s3_key":   screener_s3_key,
        "pipeline":          pipeline_stages,
    }
    try:
        existing = (
            supabase.table("screened_emails")
            .select("id")
            .eq("user_id", user_id)
            .eq("gmail_message_id", gmail_message_id)
            .limit(1)
            .execute()
        )
        if existing.data:
            supabase.table("screened_emails").update(update).eq("id", existing.data[0]["id"]).execute()
    except Exception as exc:
        log.warning("screened_email_complete_failed", gmail_message_id=gmail_message_id, error=str(exc))


def _advance_screened_email_stage(
    supabase: Any,
    user_id: str,
    gmail_message_id: str,
    new_status: str,
    completing: str,
    next_stage: str | None = None,
) -> None:
    """Mark `completing` stage as completed and optionally start `next_stage`.
    Updates processing_status to `new_status`."""
    now = datetime.now(timezone.utc).isoformat()
    try:
        existing = (
            supabase.table("screened_emails")
            .select("id, pipeline")
            .eq("user_id", user_id)
            .eq("gmail_message_id", gmail_message_id)
            .limit(1)
            .execute()
        )
        if not existing.data:
            return
        row_id = existing.data[0]["id"]
        pipeline: list[dict] = list(existing.data[0].get("pipeline") or [])
        updated = []
        for s in pipeline:
            s = dict(s)
            if s["stage"] == completing:
                s["status"] = "completed"
                s["finished_at"] = now
            elif next_stage and s["stage"] == next_stage:
                s["status"] = "in_progress"
                s["started_at"] = now
            updated.append(s)
        supabase.table("screened_emails").update(
            {"pipeline": updated, "processing_status": new_status}
        ).eq("id", row_id).execute()
    except Exception as exc:
        log.warning("advance_stage_failed", completing=completing, error=str(exc))


def _fail_screened_email(
    supabase: Any,
    user_id: str,
    gmail_message_id: str,
    error_msg: str,
) -> None:
    """Mark the screened_emails row as failed, flagging the in-progress stage."""
    now = datetime.now(timezone.utc).isoformat()
    try:
        existing = (
            supabase.table("screened_emails")
            .select("id, pipeline")
            .eq("user_id", user_id)
            .eq("gmail_message_id", gmail_message_id)
            .limit(1)
            .execute()
        )
        if not existing.data:
            return
        row_id = existing.data[0]["id"]
        pipeline: list[dict] = list(existing.data[0].get("pipeline") or [])
        updated = []
        for s in pipeline:
            s = dict(s)
            if s.get("status") == "in_progress":
                s["status"] = "failed"
                s["finished_at"] = now
                s["detail"] = error_msg[:200]
            updated.append(s)
        supabase.table("screened_emails").update(
            {"pipeline": updated, "processing_status": "failed"}
        ).eq("id", row_id).execute()
    except Exception as exc:
        log.warning("fail_screened_email_failed", error=str(exc))


def _sync_cached_screened_email(
    supabase: Any,
    user_id: str,
    gmail_message_id: str,
    deal_id: str,
) -> None:
    """When a cached deal is found, update the screened_emails stub (created by the
    endpoint before the background task ran) to reflect the completed state."""
    try:
        deal_resp = (
            supabase.table("deals")
            .select("property_name, screener_storage_path, pipeline")
            .eq("id", deal_id)
            .limit(1)
            .execute()
        )
        if not deal_resp.data:
            return
        d = deal_resp.data[0]
        _complete_screened_email(
            supabase,
            user_id,
            gmail_message_id,
            deal_id,
            screened_title=d.get("property_name"),
            screener_s3_key=d.get("screener_storage_path"),
            pipeline_stages=d.get("pipeline") or [],
        )
    except Exception as exc:
        log.warning("sync_cached_screened_email_failed", error=str(exc))


def _check_cached_result(supabase, user_id: str, gmail_message_id: str) -> str | None:
    """Return deal_id if this email was already processed, else None."""
    try:
        resp = (
            supabase.table("emails")
            .select("deal_id")
            .eq("user_id", user_id)
            .eq("gmail_message_id", gmail_message_id)
            .not_.is_("deal_id", "null")
            .limit(1)
            .execute()
        )
        if resp.data:
            return str(resp.data[0]["deal_id"])
    except Exception as exc:
        log.warning("cache_check_failed", error=str(exc))
    return None


def _persist_to_supabase(
    supabase,
    user_id: str,
    deal_id: str,
    gmail_message_id: str,
    result: dict[str, Any],
    screener_s3_key: str | None,
    s3_attachment_keys: list[dict],
    email_detail: Any,
) -> None:
    prop = result.get("property_info") or {}
    summary = result.get("deal_summary") or {}
    screening = result.get("screening_result") or {}
    now = datetime.now(timezone.utc).isoformat()

    pipeline_stages = [
        {"stage": "email_received",       "status": "completed", "finished_at": now},
        {"stage": "parsing_attachments",  "status": "completed", "finished_at": now},
        {"stage": "extracting_financials","status": "completed", "finished_at": now},
        {"stage": "running_screener",     "status": "completed", "finished_at": now},
        {"stage": "complete",             "status": "completed", "finished_at": now},
    ]

    # Write the final email row — SELECT+INSERT/UPDATE avoids ON CONFLICT constraint issues
    email_row: dict[str, Any] = {
        "user_id":          user_id,
        "gmail_message_id": gmail_message_id,
        "sender":           email_detail.sender or "Unknown",
        "subject":          email_detail.subject,
        "body_text":        email_detail.body_text,
        "received_at":      (
            email_detail.received_at.isoformat()
            if email_detail.received_at else now
        ),
        "status":           "processed",
        "attachments":      [
            {"filename": k["filename"], "s3_key": k["s3_key"], "type": k["type"]}
            for k in s3_attachment_keys
        ],
        "deal_id":          deal_id,
    }
    internal_email_id = _safe_email_write(supabase, user_id, gmail_message_id, email_row)

    # Insert deal row
    deal_row: dict[str, Any] = {
        "id":                   deal_id,
        "user_id":              user_id,
        "email_id":             internal_email_id,
        "property_name":        prop.get("property_name", "Unknown Property"),
        "asset_class":          prop.get("property_type"),
        "recommendation":       screening.get("recommendation"),
        "confidence":           screening.get("confidence"),
        "risk_rating":          screening.get("risk_rating"),
        "screener_storage_path":screener_s3_key,
        "email_draft":          _build_email_draft(result),
        "metrics":              screening.get("key_metrics", []),
        "key_metrics":          screening.get("key_metrics", []),
        "highlights": [
            {
                "title":  "Investment Highlights",
                "detail": summary.get("investment_highlights", ""),
            }
        ],
        "risks": [
            {
                "title":    "Investment Risks & Underwriting Flags",
                "detail":   summary.get("investment_risks", ""),
                "severity": screening.get("risk_rating", "moderate"),
            }
        ],
        "pipeline":         pipeline_stages,
        "financial_summary":screening.get("financial_summary"),
        "sources_and_uses": screening.get("sources_and_uses"),
        "sponsor_overview": summary.get("sponsor_overview"),
        "location_summary": summary.get("location_summary"),
        "property_info": result.get("property_info") or {},
        "financials":    {
            "historical": result.get("financials") or {},
            "proforma":   result.get("proforma") or {},
        },
        "s3_folder_key": gmail_message_id,
        "created_at":    now,
        "updated_at":    now,
    }
    try:
        supabase.table("deals").insert(deal_row).execute()
    except Exception as exc:
        log.error("deal_insert_failed", deal_id=deal_id, error=str(exc))
        raise

    _complete_screened_email(
        supabase,
        user_id,
        gmail_message_id,
        deal_id,
        screened_title=prop.get("property_name"),
        screener_s3_key=screener_s3_key,
        pipeline_stages=pipeline_stages,
    )


def _build_email_draft(result: dict[str, Any]) -> str:
    prop = result.get("property_info") or {}
    summary = result.get("deal_summary") or {}
    screening = result.get("screening_result") or {}
    loan = result.get("loan_terms") or {}

    rec = (screening.get("recommendation") or "negotiate").upper()
    rec_map = {"PROCEED": "Proceed to Diligence", "NEGOTIATE": "Negotiate", "PASS": "Pass"}
    rec_label = rec_map.get(rec, rec)

    metrics_lines = "\n".join(
        f"  {m['label']}: {m['value']}"
        for m in (screening.get("key_metrics") or [])
    )

    return (
        f"Team,\n\n"
        f"Please find below the screening summary for {prop.get('property_name', '[Property]')}, "
        f"a {prop.get('property_type', 'senior housing')} community in {prop.get('city_state', '[Location]')}.\n\n"
        f"KEY METRICS:\n{metrics_lines}\n\n"
        f"HIGHLIGHTS:\n{summary.get('investment_highlights', '')}\n\n"
        f"RISKS & FLAGS:\n{summary.get('investment_risks', '')}\n\n"
        f"Recommendation: {rec_label}\n\n"
        f"— Bloomfield Capital Originations"
    )


# ─────────────────────────────────────────────────────────────────────────────
#  AWS S3 HELPERS  (synchronous — must be called via run_in_executor)
# ─────────────────────────────────────────────────────────────────────────────

def _get_s3_client(settings: Any) -> Any:
    return boto3.client(
        "s3",
        aws_access_key_id=settings.aws_access_key_id,
        aws_secret_access_key=settings.aws_secret_access_key,
        region_name=settings.aws_region,
    )


def _upload_attachments_sync(
    gmail_message_id: str,
    attachments: list[dict[str, Any]],
    settings: Any,
) -> list[dict[str, Any]]:
    s3 = _get_s3_client(settings)
    uploaded = []
    for att in attachments:
        key = f"{gmail_message_id}/{att['filename']}"
        try:
            s3.put_object(
                Bucket=settings.aws_s3_bucket,
                Key=key,
                Body=att["data"],
                ContentType="application/octet-stream",
            )
            uploaded.append({"filename": att["filename"], "s3_key": key, "type": att["type"]})
        except Exception as exc:
            log.warning("s3_upload_failed", key=key, error=str(exc))
    return uploaded


def _s3_put_bytes(key: str, data: bytes, content_type: str, settings: Any) -> None:
    s3 = _get_s3_client(settings)
    s3.put_object(Bucket=settings.aws_s3_bucket, Key=key, Body=data, ContentType=content_type)


def _s3_presigned_url(key: str, settings: Any, expires: int = 3600) -> str:
    s3 = _get_s3_client(settings)
    return s3.generate_presigned_url(
        "get_object",
        Params={"Bucket": settings.aws_s3_bucket, "Key": key},
        ExpiresIn=expires,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  DOCUMENT TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def _build_document_text_blocks(attachments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Extract text from PDFs and Excel files and return as Claude text content blocks."""
    blocks: list[dict[str, Any]] = []
    for att in attachments:
        if att["type"] == "pdf":
            text = _extract_pdf_text(att["data"], att["filename"])
        elif att["type"] == "excel":
            text = _extract_excel_text(att["data"], att["filename"])
        else:
            continue
        if text.strip():
            blocks.append({
                "type": "text",
                "text": (
                    f"=== DOCUMENT: {att['filename']} ===\n"
                    f"{text[:60_000]}\n"   # guard against huge docs
                    f"=== END: {att['filename']} ==="
                ),
            })
    return blocks


def _extract_pdf_text(data: bytes, filename: str) -> str:
    try:
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            pages: list[str] = []
            for i, page in enumerate(pdf.pages):
                tables = page.extract_tables()
                if tables:
                    table_text = "\n\n".join(
                        "\n".join(
                            "\t".join(str(c) if c is not None else "" for c in row)
                            for row in tbl
                        )
                        for tbl in tables
                    )
                    plain = page.extract_text() or ""
                    pages.append(f"--- Page {i + 1} ---\n{plain}\n\n[TABLES]\n{table_text}")
                else:
                    pages.append(f"--- Page {i + 1} ---\n{page.extract_text() or ''}")
        return "\n\n".join(pages)
    except Exception as exc:
        log.warning("pdf_extract_failed", filename=filename, error=str(exc))
        return ""


def _extract_excel_text(data: bytes, filename: str) -> str:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        sheets: list[str] = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                line = "\t".join(cells).rstrip()
                if line.replace("\t", "").strip():
                    rows.append(line)
            if rows:
                sheets.append(f"--- Sheet: {name} ---\n" + "\n".join(rows))
        return "\n\n".join(sheets)
    except Exception as exc:
        log.warning("excel_extract_failed", filename=filename, error=str(exc))
        return ""


# ─────────────────────────────────────────────────────────────────────────────
#  CLAUDE API CALL
# ─────────────────────────────────────────────────────────────────────────────

async def _call_claude(
    email_detail: Any,
    doc_blocks: list[dict[str, Any]],
    settings: Any,
    additional_instructions: str | None = None,
) -> dict[str, Any]:
    client = get_claude_client()

    content: list[dict[str, Any]] = []

    # Email header + body
    content.append({
        "type": "text",
        "text": (
            f"FROM: {email_detail.sender} <{email_detail.sender_email or ''}>\n"
            f"SUBJECT: {email_detail.subject}\n"
            f"DATE: {email_detail.received_at}\n\n"
            f"EMAIL BODY:\n"
            f"{email_detail.body_text or '(No plain text — HTML-only email)'}\n"
        ),
    })

    # Extracted document content
    content.extend(doc_blocks)

    if not doc_blocks:
        content.append({
            "type": "text",
            "text": (
                "[NOTE: No attachments were provided with this email. "
                "Analyze based on the email body only and flag all missing "
                "financial documents as underwriting gaps.]"
            ),
        })

    # Optional analyst instructions — injected immediately before the task
    # directive so Claude treats them as high-priority guidance.
    if additional_instructions and additional_instructions.strip():
        content.append({
            "type": "text",
            "text": (
                "════════════════════════════════════════════════════════\n"
                "SPECIAL INSTRUCTIONS FROM THE ANALYST (HIGHEST PRIORITY — MUST FOLLOW):\n"
                "════════════════════════════════════════════════════════\n"
                f"{additional_instructions.strip()}\n"
                "════════════════════════════════════════════════════════\n"
                "Apply the above instructions when producing your analysis and output."
            ),
        })

    # Final task instruction
    content.append({
        "type": "text",
        "text": (
            "Analyze the above email and all provided document extracts. "
            "Apply the senior housing underwriting framework to extract every "
            "available data point, compute the required metrics, write the three "
            "Deal Summary narrative sections, then call submit_screening_result "
            "with the complete structured output. Flag all risks explicitly and "
            "note every missing data item."
        ),
    })

    response = await client.messages.create(
        model=settings.claude_model,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": content}],
        tools=[SCREENER_TOOL],
        tool_choice={"type": "tool", "name": "submit_screening_result"},
        max_tokens=8192,
    )

    for block in response.content:
        if hasattr(block, "type") and block.type == "tool_use":
            if block.name == "submit_screening_result":
                return dict(block.input)

    raise ValueError(
        "Claude did not call submit_screening_result. "
        f"Stop reason: {response.stop_reason}. "
        f"Content types: {[getattr(b, 'type', '?') for b in response.content]}"
    )


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL SCREENER GENERATION
# ─────────────────────────────────────────────────────────────────────────────

# Column letter → openpyxl column index mapping for financial periods
_FIN_COL_MAP = {
    "ye_2022":         "C",
    "ye_2023":         "D",
    "ye_2024":         "E",
    "t11m_annualized": "F",
    "t3m_annualized":  "G",
}
_PROFORMA_COL_MAP = {
    "year_1": "H",
    "year_2": "J",
    "year_3": "L",
}
# Row positions for each P&L line item (Summary sheet, Bloomfield template)
_FIN_ROW_MAP = {
    "il_revenue":                 81,
    "al_revenue":                 82,
    "mc_revenue":                 83,
    "community_fee":              85,
    "additional_occupant_income": 86,
    "other_income":               87,
    "salaries_wages":             91,
    "taxes_benefits":             92,
    "contract_labor":             93,
    "general_administrative":     94,
    "plant_operations":           95,
    "insurance":                  96,
    "real_estate_taxes":          97,
    "housekeeping_laundry":       98,
    "activities_social_services": 99,
    "resident_care":             100,
    "utilities_fixed":           101,
    "utilities_variable":        102,
    "dietary":                   103,
    "food_costs":                104,
    "marketing":                 105,
    "bad_debt":                  106,
    "commissions":               107,
    "management_fee":            108,
}
_UNIT_ROW_MAP = {
    "AL Studio":    48,
    "AL 1-Bed":     49,
    "AL 2-Bed":     50,
    "MC Companion": 51,
    "MC Studio":    52,
    "MC 1-Bed":     53,
}


def _generate_excel_screener(result: dict[str, Any], template_path: str) -> bytes:
    """Populate the Bloomfield screener template (or a blank workbook) and
    inject Deal Summary text boxes. Returns the Excel file as bytes."""
    tmpl = Path(template_path)
    log.info("screener_template_path", resolved=str(tmpl.resolve()))
    if tmpl.exists():
        wb = openpyxl.load_workbook(tmpl)
        log.info("screener_template_sheets", sheets=wb.sheetnames)
        ws = wb["Summary"] if "Summary" in wb.sheetnames else wb.active
    else:
        log.warning("screener_template_missing", path=str(tmpl))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"  # type: ignore[assignment]

    prop   = result.get("property_info") or {}
    loan   = result.get("loan_terms") or {}
    pf     = result.get("proforma") or {}
    fins   = result.get("financials") or {}

    # ── HEADER ──────────────────────────────────────────────────────────────
    ws["C3"] = prop.get("property_name")
    ws["C4"] = prop.get("address")
    ws["C5"] = prop.get("city_state")
    ws["M3"] = prop.get("deal_type")
    ws["M4"] = prop.get("property_type")
    ws["N5"] = prop.get("total_units")

    # ── VALUE SUMMARY ────────────────────────────────────────────────────────
    ws["F30"] = loan.get("requested_loan_amount")
    ws["F32"] = loan.get("purchase_price")
    ws["I31"] = loan.get("interest_rate", 0.11)
    for cell_row, yr_key in [("34", "year_1"), ("35", "year_2")]:
        yr = pf.get(yr_key) or {}
        noi = yr.get("noi")
        cap = yr.get("cap_rate", 0.09)
        ws[f"D{cell_row}"] = noi
        ws[f"E{cell_row}"] = cap
        if noi and cap:
            ws[f"F{cell_row}"] = noi / cap

    # ── SUBJECT PROPERTY DETAILS ─────────────────────────────────────────────
    ws["C40"] = prop.get("deal_type")
    ws["C41"] = prop.get("property_name")
    ws["C42"] = prop.get("address")
    ws["C43"] = prop.get("city_state")
    ws["C44"] = prop.get("county")
    ws["C45"] = prop.get("property_type")
    ws["C47"] = prop.get("al_units")
    ws["C48"] = prop.get("mc_units")
    ws["C49"] = prop.get("total_units")
    ws["C50"] = prop.get("total_beds")
    ws["C51"] = prop.get("bed_occupancy_pct")
    ws["C53"] = prop.get("year_built")
    ws["D54"] = prop.get("building_sf")

    # ── RENT ROLL ────────────────────────────────────────────────────────────
    for unit in result.get("unit_mix") or []:
        row = _UNIT_ROW_MAP.get(unit.get("unit_type", ""))
        if not row:
            continue
        units = unit.get("total_units") or 0
        occ   = unit.get("occupancy_pct") or 0
        rent  = unit.get("rent_per_resident_mo") or 0
        care  = unit.get("care_per_resident_mo") or 0
        ws.cell(row=row, column=7).value  = units
        ws.cell(row=row, column=8).value  = unit.get("total_beds")
        ws.cell(row=row, column=9).value  = occ
        ws.cell(row=row, column=10).value = rent
        ws.cell(row=row, column=11).value = care
        ws.cell(row=row, column=12).value = rent + care
        ws.cell(row=row, column=13).value = units * occ * rent
        ws.cell(row=row, column=14).value = units * occ * (rent + care)

    # ── LOAN SOURCES & USES ──────────────────────────────────────────────────
    ws["N61"] = loan.get("interest_reserve_months", 12)
    ws["J62"] = loan.get("capex_reserve")
    ws["N63"] = loan.get("bloomfield_points_pct", 0.02)
    ws["N64"] = loan.get("broker_points_pct", 0.01)

    # ── FINANCIAL SUMMARY ────────────────────────────────────────────────────
    for period_key, col_letter in _FIN_COL_MAP.items():
        period_data = fins.get(period_key)
        if not period_data:
            continue
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        for field, row in _FIN_ROW_MAP.items():
            val = period_data.get(field)
            if val is not None:
                ws.cell(row=row, column=col_idx).value = val

    for yr_key, col_letter in _PROFORMA_COL_MAP.items():
        yr = pf.get(yr_key) or {}
        noi = yr.get("noi")
        if noi is not None:
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            ws.cell(row=111, column=col_idx).value = noi

    deal_summary = result.get("deal_summary") or {}
    overview   = deal_summary.get("property_overview", "")
    highlights = deal_summary.get("investment_highlights", "")
    risks      = deal_summary.get("investment_risks", "")

    from openpyxl.styles import Alignment
    wrap = Alignment(wrap_text=True, vertical="top")

    combined = f"{overview}\n\n{highlights}".strip()
    ws["B8"] = combined
    ws["B8"].alignment = wrap
    for r in range(8, 12):
        ws.row_dimensions[r].height = None

    ws["B13"] = risks
    ws["B13"].alignment = wrap
    for r in range(13, 26):
        ws.row_dimensions[r].height = None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
#  SCREENER DOWNLOAD HELPER  (called by the deals route)
# ─────────────────────────────────────────────────────────────────────────────

async def get_screener_presigned_url(s3_key: str) -> str:
    """Return a 1-hour pre-signed S3 URL for downloading the screener Excel."""
    settings = get_settings()
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _s3_presigned_url, s3_key, settings)
